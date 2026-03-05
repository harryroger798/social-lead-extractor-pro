"""Export leads to Excel, CSV, JSON, and HTML formats."""
import csv
import json
import io
import os
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


EXPORT_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "exports")


def ensure_export_dir() -> str:
    os.makedirs(EXPORT_DIR, exist_ok=True)
    return EXPORT_DIR


def _get_filename(format_type: str, session_name: str = "export") -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = "".join(c if c.isalnum() or c in "-_ " else "" for c in session_name).strip()
    safe_name = safe_name.replace(" ", "_")[:50]
    return f"{safe_name}_{timestamp}.{format_type}"


def export_to_csv(leads: list[dict], session_name: str = "export") -> str:
    """Export leads to CSV file. Returns file path."""
    ensure_export_dir()
    filename = _get_filename("csv", session_name)
    filepath = os.path.join(EXPORT_DIR, filename)

    if not leads:
        return filepath

    fieldnames = ["email", "phone", "name", "platform", "source_url",
                  "keyword", "country", "email_type", "verified", "quality_score",
                  "extracted_at"]

    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for lead in leads:
            row = {k: lead.get(k, "") for k in fieldnames}
            row["verified"] = "Yes" if lead.get("verified") else "No"
            writer.writerow(row)

    return filepath


def export_to_json(leads: list[dict], session_name: str = "export") -> str:
    """Export leads to JSON file. Returns file path."""
    ensure_export_dir()
    filename = _get_filename("json", session_name)
    filepath = os.path.join(EXPORT_DIR, filename)

    export_data = {
        "exported_at": datetime.now().isoformat(),
        "total_leads": len(leads),
        "leads": leads,
    }

    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(export_data, f, indent=2, ensure_ascii=False)

    return filepath


def export_to_excel(leads: list[dict], session_name: str = "export") -> str:
    """Export leads to Excel file. Returns file path."""
    ensure_export_dir()
    filename = _get_filename("xlsx", session_name)
    filepath = os.path.join(EXPORT_DIR, filename)

    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    ws.title = "Leads"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["Email", "Phone", "Name", "Platform", "Source URL",
               "Keyword", "Country", "Email Type", "Verified", "Quality Score",
               "Extracted At"]
    field_keys = ["email", "phone", "name", "platform", "source_url",
                  "keyword", "country", "email_type", "verified", "quality_score",
                  "extracted_at"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, lead in enumerate(leads, 2):
        for col_idx, key in enumerate(field_keys, 1):
            value = lead.get(key, "")
            if key == "verified":
                value = "Yes" if value else "No"
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    # Auto-width columns
    for col in ws.columns:
        max_length = 0
        col_letter = None
        for cell in col:
            if col_letter is None:
                col_letter = cell.column_letter
            try:
                cell_len = len(str(cell.value or ""))
                if cell_len > max_length:
                    max_length = cell_len
            except Exception:
                pass
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    wb.save(filepath)
    return filepath


def export_to_html(leads: list[dict], session_name: str = "export") -> str:
    """Export leads to HTML file. Returns file path."""
    ensure_export_dir()
    filename = _get_filename("html", session_name)
    filepath = os.path.join(EXPORT_DIR, filename)

    html_parts = [
        "<!DOCTYPE html>",
        "<html><head>",
        "<meta charset='utf-8'>",
        f"<title>{session_name} - Lead Export</title>",
        "<style>",
        "body{font-family:system-ui,-apple-system,sans-serif;background:#0F172A;color:#F8FAFC;padding:20px;}",
        "h1{color:#3B82F6;margin-bottom:5px;}",
        "p.meta{color:#94A3B8;font-size:14px;margin-bottom:20px;}",
        "table{width:100%;border-collapse:collapse;background:#1E293B;border-radius:8px;overflow:hidden;}",
        "th{background:#3B82F6;color:white;padding:12px 16px;text-align:left;font-size:13px;}",
        "td{padding:10px 16px;border-bottom:1px solid #334155;font-size:13px;}",
        "tr:hover{background:#273549;}",
        ".verified{color:#22C55E;}.unverified{color:#EF4444;}",
        ".score-high{color:#22C55E;font-weight:bold;}",
        ".score-mid{color:#F59E0B;font-weight:bold;}",
        ".score-low{color:#EF4444;font-weight:bold;}",
        "</style></head><body>",
        f"<h1>{session_name}</h1>",
        f"<p class='meta'>Exported {len(leads)} leads on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>",
        "<table><thead><tr>",
        "<th>Email</th><th>Phone</th><th>Name</th><th>Platform</th>",
        "<th>Keyword</th><th>Country</th><th>Type</th><th>Verified</th><th>Score</th>",
        "</tr></thead><tbody>",
    ]

    for lead in leads:
        verified = lead.get("verified", False)
        score = lead.get("quality_score", 0)
        score_class = "score-high" if score >= 80 else "score-mid" if score >= 60 else "score-low"
        verified_class = "verified" if verified else "unverified"
        verified_text = "Yes" if verified else "No"

        html_parts.append("<tr>")
        html_parts.append(f"<td>{lead.get('email', '')}</td>")
        html_parts.append(f"<td>{lead.get('phone', '')}</td>")
        html_parts.append(f"<td>{lead.get('name', '')}</td>")
        html_parts.append(f"<td>{lead.get('platform', '')}</td>")
        html_parts.append(f"<td>{lead.get('keyword', '')}</td>")
        html_parts.append(f"<td>{lead.get('country', '')}</td>")
        html_parts.append(f"<td>{lead.get('email_type', '')}</td>")
        html_parts.append(f"<td class='{verified_class}'>{verified_text}</td>")
        html_parts.append(f"<td class='{score_class}'>{score}</td>")
        html_parts.append("</tr>")

    html_parts.append("</tbody></table></body></html>")

    with open(filepath, "w", encoding="utf-8") as f:
        f.write("\n".join(html_parts))

    return filepath


def export_leads(
    leads: list[dict],
    format_type: str = "csv",
    session_name: str = "export",
) -> str:
    """Export leads in the specified format. Returns file path."""
    exporters = {
        "csv": export_to_csv,
        "json": export_to_json,
        "excel": export_to_excel,
        "xlsx": export_to_excel,
        "html": export_to_html,
    }
    exporter = exporters.get(format_type, export_to_csv)
    return exporter(leads, session_name)


def export_leads_bytes(
    leads: list[dict],
    format_type: str = "csv",
) -> tuple[bytes, str, str]:
    """Export leads and return bytes, content_type, and filename."""
    if format_type in ("excel", "xlsx"):
        filepath = export_to_excel(leads)
        with open(filepath, "rb") as f:
            data = f.read()
        os.unlink(filepath)
        return data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "leads.xlsx"

    if format_type == "json":
        output = json.dumps({"leads": leads}, indent=2).encode("utf-8")
        return output, "application/json", "leads.json"

    if format_type == "html":
        filepath = export_to_html(leads)
        with open(filepath, "rb") as f:
            data = f.read()
        os.unlink(filepath)
        return data, "text/html", "leads.html"

    # Default CSV
    output = io.StringIO()
    if leads:
        fieldnames = ["email", "phone", "name", "platform", "source_url",
                      "keyword", "country", "email_type", "verified", "quality_score"]
        writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for lead in leads:
            writer.writerow(lead)
    return output.getvalue().encode("utf-8"), "text/csv", "leads.csv"
